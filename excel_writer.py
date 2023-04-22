import openpyxl
import os


def write_dict_to_file(filename, sheetname, data):
    filename = f'{filename}.xlsx'
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()

    if sheetname in workbook.worksheets:
        worksheet = workbook[sheetname]
    else:
        worksheet = workbook.create_sheet(sheetname)
        i = 1
        for key in data.keys():
            worksheet.cell(row=1, column=i, value=key)
            i += 1

    new_row = worksheet.max_row + 1

    i = 1
    for key in data.keys():
        worksheet.cell(row=new_row, column=i, value=data[key])
        i += 1

    workbook.save(filename)